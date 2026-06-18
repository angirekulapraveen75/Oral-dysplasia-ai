"""
Generate a professional Excel DAST Report from report.json
"""
import json
import os
import sys
import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORT_JSON = os.path.join(SCRIPT_DIR, "report.json")
OUTPUT_XLSX = os.path.join(SCRIPT_DIR, "DAST_Security_Report.xlsx")

# ── Styles ──────────────────────────────────────────────────────────
DARK_BG = PatternFill("solid", fgColor="1B1F3B")
HEADER_BG = PatternFill("solid", fgColor="2D336B")
CRITICAL_BG = PatternFill("solid", fgColor="D32F2F")
HIGH_BG = PatternFill("solid", fgColor="E65100")
MEDIUM_BG = PatternFill("solid", fgColor="F9A825")
LOW_BG = PatternFill("solid", fgColor="1565C0")
INFO_BG = PatternFill("solid", fgColor="2E7D32")
PASS_BG = PatternFill("solid", fgColor="1B5E20")
FAIL_BG = PatternFill("solid", fgColor="B71C1C")
WHITE_BG = PatternFill("solid", fgColor="FFFFFF")
LIGHT_GRAY_BG = PatternFill("solid", fgColor="F5F5F5")
LIGHT_BLUE_BG = PatternFill("solid", fgColor="E3F2FD")

TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="1B1F3B")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=False, color="555555")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10, color="333333")
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
LINK_FONT = Font(name="Calibri", size=10, color="1565C0", underline="single")
SEV_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
KPI_NUM_FONT = Font(name="Calibri", size=28, bold=True, color="1B1F3B")
KPI_LABEL_FONT = Font(name="Calibri", size=10, bold=False, color="777777")

THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def sev_fill(severity):
    s = severity.upper() if severity else "INFO"
    if s == "CRITICAL": return CRITICAL_BG
    if s == "HIGH": return HIGH_BG
    if s == "MEDIUM": return MEDIUM_BG
    if s == "LOW": return LOW_BG
    return INFO_BG


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def body_cell(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or BODY_FONT
    cell.fill = fill or (LIGHT_GRAY_BG if row % 2 == 0 else WHITE_BG)
    cell.alignment = align or LEFT_WRAP
    cell.border = THIN_BORDER
    return cell


# ── Load data ───────────────────────────────────────────────────────
with open(REPORT_JSON, "r") as f:
    results = json.load(f)

findings = [r for r in results if r.get("finding")]
total = len(results)
total_findings = len(findings)
by_sev = {}
for r in findings:
    s = r["severity"]
    by_sev.setdefault(s, []).append(r)

by_cat = {}
for r in results:
    c = r["test_category"]
    by_cat.setdefault(c, []).append(r)

# ── Create workbook ─────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Executive Summary"
ws1.sheet_properties.tabColor = "1B1F3B"
set_col_widths(ws1, [3, 22, 22, 22, 22, 22, 3])

# Title block
ws1.merge_cells("B2:F2")
c = ws1["B2"]
c.value = "DAST Security Assessment Report"
c.font = TITLE_FONT
c.alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("B3:F3")
c = ws1["B3"]
c.value = "OralDysplasia AI — API Security Testing"
c.font = SUBTITLE_FONT

ws1.merge_cells("B4:F4")
c = ws1["B4"]
c.value = f"Target: https://oral-dysplasia-ai.vercel.app  |  Date: {datetime.date.today().strftime('%d %B %Y')}  |  Tester: Automated DAST Runner"
c.font = Font(name="Calibri", size=9, color="888888")

# KPI row
kpi_data = [
    ("Total Tests", str(total)),
    ("Findings", str(total_findings)),
    ("CRITICAL", str(len(by_sev.get("CRITICAL", [])))),
    ("HIGH", str(len(by_sev.get("HIGH", [])))),
    ("MEDIUM", str(len(by_sev.get("MEDIUM", [])))),
]
kpi_fills = [LIGHT_BLUE_BG, FAIL_BG if total_findings > 0 else PASS_BG, CRITICAL_BG, HIGH_BG, MEDIUM_BG]
kpi_num_colors = ["1B1F3B", "FFFFFF", "FFFFFF", "FFFFFF", "000000"]
kpi_label_colors = ["777777", "FFFFFF", "FFFFFF", "FFFFFF", "333333"]

for i, (label, val) in enumerate(kpi_data):
    col = i + 2
    cell_num = ws1.cell(row=6, column=col, value=val)
    cell_num.font = Font(name="Calibri", size=28, bold=True, color=kpi_num_colors[i])
    cell_num.fill = kpi_fills[i]
    cell_num.alignment = CENTER
    cell_num.border = THIN_BORDER

    cell_label = ws1.cell(row=7, column=col, value=label)
    cell_label.font = Font(name="Calibri", size=10, bold=False, color=kpi_label_colors[i])
    cell_label.fill = kpi_fills[i]
    cell_label.alignment = CENTER
    cell_label.border = THIN_BORDER

ws1.row_dimensions[6].height = 45
ws1.row_dimensions[7].height = 22

# Category breakdown table
header_row(ws1, 9, ["", "Test Category", "Tests Run", "Passed", "Failed", ""])
categories_display = [
    ("AuthN Bypass (No Token)", "authn_bypass"),
    ("AuthN Bypass (Bad Tokens)", "authn_bypass"),
    ("AuthZ / RBAC Matrix", "authz_rbac"),
    ("IDOR", "idor"),
    ("Token Tampering", "token_tampering"),
    ("Injection Probes", "injection"),
    ("Rate Limiting", "rate_limiting"),
    ("Hardcoded Credentials", "hardcoded_creds"),
    ("Role Self-Assignment", "authz_rbac"),
    ("CORS Misconfiguration", "cors"),
]

# Recompute per-category
cat_stats = {}
for r in results:
    cat = r["test_category"]
    cat_stats.setdefault(cat, {"total": 0, "findings": 0})
    cat_stats[cat]["total"] += 1
    if r.get("finding"):
        cat_stats[cat]["findings"] += 1

row = 10
for cat_key, stats in cat_stats.items():
    display_name = cat_key.replace("_", " ").title()
    t = stats["total"]
    f = stats["findings"]
    p = t - f
    body_cell(ws1, row, 2, display_name, font=BOLD_FONT)
    body_cell(ws1, row, 3, t, align=CENTER)
    body_cell(ws1, row, 4, p, align=CENTER, fill=PASS_BG if p == t else None,
              font=SEV_FONT if p == t else BODY_FONT)
    body_cell(ws1, row, 5, f, align=CENTER, fill=FAIL_BG if f > 0 else None,
              font=SEV_FONT if f > 0 else BODY_FONT)
    row += 1

# Totals
body_cell(ws1, row, 2, "TOTAL", font=Font(name="Calibri", size=11, bold=True, color="1B1F3B"))
body_cell(ws1, row, 3, total, font=BOLD_FONT, align=CENTER)
body_cell(ws1, row, 4, total - total_findings, font=BOLD_FONT, align=CENTER)
body_cell(ws1, row, 5, total_findings, font=BOLD_FONT, align=CENTER,
          fill=FAIL_BG if total_findings > 0 else PASS_BG)

# ════════════════════════════════════════════════════════════════════
# SHEET 2: Findings
# ════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Findings")
ws2.sheet_properties.tabColor = "D32F2F"
set_col_widths(ws2, [5, 12, 12, 40, 12, 25, 65])

ws2.merge_cells("B1:G1")
ws2["B1"].value = "Security Findings"
ws2["B1"].font = TITLE_FONT

headers = ["#", "Severity", "Category", "Endpoint", "Method", "Role", "Description / Note"]
header_row(ws2, 3, [""] + headers)

# Sort findings by severity
sev_order = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "info": 4}
sorted_findings = sorted(findings, key=lambda x: sev_order.get(x["severity"], 5))

row = 4
for i, f in enumerate(sorted_findings, 1):
    body_cell(ws2, row, 2, i, align=CENTER)
    sev_cell = body_cell(ws2, row, 3, f["severity"].upper(), font=SEV_FONT, fill=sev_fill(f["severity"]), align=CENTER)
    body_cell(ws2, row, 4, f["test_category"].replace("_", " ").title())
    body_cell(ws2, row, 5, f["endpoint"])
    body_cell(ws2, row, 6, f["method"], align=CENTER)
    body_cell(ws2, row, 7, f.get("role", ""))
    # Note in next merged row
    row += 1
    ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
    note_cell = body_cell(ws2, row, 4, f["note"], font=Font(name="Calibri", size=9, italic=True, color="555555"),
                          align=LEFT_TOP)
    ws2.row_dimensions[row].height = 35
    row += 1

# Add remediation section
row += 2
ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
ws2.cell(row=row, column=2, value="Remediation Recommendations").font = Font(name="Calibri", size=14, bold=True, color="1B1F3B")
row += 1

remediations = [
    ("CRITICAL", "F1: Hardcoded SECRET_KEY",
     "Remove default value from config.py. Require SECRET_KEY via environment variable. Add startup check that blocks if not set. Rotate the current key immediately."),
    ("HIGH", "F2: Hardcoded Encryption Key",
     "Generate a real Fernet key, store in environment variable only, never commit to source code."),
    ("HIGH", "F3: Hardcoded DB Credentials",
     "Remove default DATABASE_URL from config.py. Use environment variable for all environments."),
    ("HIGH", "F4: User-Controlled Role on Signup",
     "Add server-side role validation: ALLOWED_ROLES = {'Consultant Pathologist', 'Resident', 'Lab Tech'}. Reject any role not in whitelist."),
    ("HIGH", "F5: No RBAC Enforcement",
     "Implement role-checking dependency. Add require_role() decorator to sensitive endpoints (analysis/run, review, export)."),
    ("MEDIUM", "F6: No Rate Limiting",
     "Add rate limiting middleware (e.g. slowapi for FastAPI). Suggested: 5 login attempts/minute/IP, 60 API calls/minute/user."),
    ("MEDIUM", "F7: CORS Wildcard",
     "Replace allow_origins=['*'] with explicit allowed origins: ['https://oral-dysplasia-ai.vercel.app', 'http://localhost:8000']."),
]

header_row(ws2, row, ["", "Severity", "Finding", "", "", "", "Recommended Fix"])
row += 1

for sev, finding, fix in remediations:
    body_cell(ws2, row, 2, "", align=CENTER)
    body_cell(ws2, row, 3, sev, font=SEV_FONT, fill=sev_fill(sev), align=CENTER)
    ws2.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    body_cell(ws2, row, 4, finding, font=BOLD_FONT)
    ws2.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
    body_cell(ws2, row, 6, fix, align=LEFT_WRAP)
    ws2.row_dimensions[row].height = 45
    row += 1


# ════════════════════════════════════════════════════════════════════
# SHEET 3: Full Test Results
# ════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("All Test Results")
ws3.sheet_properties.tabColor = "1565C0"
set_col_widths(ws3, [5, 40, 10, 22, 10, 16, 10, 12, 15, 50, 22])

ws3.merge_cells("B1:K1")
ws3["B1"].value = "Complete Test Results Matrix"
ws3["B1"].font = TITLE_FONT

headers3 = ["#", "Endpoint", "Method", "Role", "Status", "Expected", "Finding?", "Severity",
            "Response (ms)", "Note", "Timestamp"]
header_row(ws3, 3, [""] + headers3)

row = 4
for i, r in enumerate(results, 1):
    is_finding = r.get("finding", False)
    body_cell(ws3, row, 2, i, align=CENTER)
    body_cell(ws3, row, 3, r["endpoint"])
    body_cell(ws3, row, 4, r["method"], align=CENTER)
    body_cell(ws3, row, 5, r.get("role", ""))
    body_cell(ws3, row, 6, str(r.get("status", "")), align=CENTER)
    body_cell(ws3, row, 7, r.get("expected_status", ""), align=CENTER)

    finding_text = "YES" if is_finding else "NO"
    finding_fill = FAIL_BG if is_finding else PASS_BG
    body_cell(ws3, row, 8, finding_text, font=SEV_FONT, fill=finding_fill, align=CENTER)

    if is_finding:
        body_cell(ws3, row, 9, r.get("severity", "").upper(), font=SEV_FONT,
                  fill=sev_fill(r.get("severity", "")), align=CENTER)
    else:
        body_cell(ws3, row, 9, "info", align=CENTER)

    body_cell(ws3, row, 10, r.get("response_time_ms", 0), align=CENTER)
    body_cell(ws3, row, 11, r.get("note", ""), align=LEFT_WRAP)
    body_cell(ws3, row, 12, r.get("timestamp", ""), font=Font(name="Calibri", size=8, color="999999"))
    row += 1

# ════════════════════════════════════════════════════════════════════
# SHEET 4: Endpoints
# ════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Endpoints Discovered")
ws4.sheet_properties.tabColor = "2E7D32"
set_col_widths(ws4, [5, 6, 10, 45, 15, 20, 30])

ws4.merge_cells("B1:G1")
ws4["B1"].value = "API Endpoints Discovered"
ws4["B1"].font = TITLE_FONT

endpoints = [
    (1, "POST", "/api/v1/auth/signup", "Public", "None", "User registration"),
    (2, "POST", "/api/v1/auth/login", "Public", "None", "User authentication"),
    (3, "POST", "/api/v1/auth/forgot-password", "Public", "None", "Password recovery"),
    (4, "POST", "/api/v1/slides/upload", "JWT Required", "None (should restrict)", "Multipart file upload"),
    (5, "GET", "/api/v1/slides/library", "JWT Required", "None", "Paginated slide list"),
    (6, "GET", "/api/v1/slides/{slide_id}", "JWT Required", "None (IDOR risk)", "Single slide detail"),
    (7, "GET", "/api/v1/slides/stats/dashboard", "JWT Required", "None", "Dashboard aggregates"),
    (8, "POST", "/api/v1/analysis/run", "JWT Required", "None (should restrict)", "Trigger AI analysis"),
    (9, "GET", "/api/v1/analysis/{slide_id}/result", "JWT Required", "None (IDOR risk)", "Get analysis results"),
    (10, "PUT", "/api/v1/analysis/{slide_id}/review", "JWT Required", "None (should restrict)", "Submit pathologist review"),
    (11, "GET", "/api/v1/reports/{slide_id}/export", "JWT Required", "None (should restrict)", "Export diagnostic report"),
    (12, "GET", "/openapi.json", "Public", "None", "OpenAPI specification"),
    (13, "GET", "/docs", "Public", "None", "Swagger UI documentation"),
]

header_row(ws4, 3, ["", "#", "Method", "Path", "Auth", "RBAC", "Description"])
for i, (num, method, path, auth, rbac, desc) in enumerate(endpoints):
    row = 4 + i
    body_cell(ws4, row, 2, num, align=CENTER)
    body_cell(ws4, row, 3, method, align=CENTER, font=BOLD_FONT)
    body_cell(ws4, row, 4, path, font=BOLD_FONT)
    body_cell(ws4, row, 5, auth, align=CENTER)
    body_cell(ws4, row, 6, rbac, align=CENTER)
    body_cell(ws4, row, 7, desc)

# ── Save ────────────────────────────────────────────────────────────
wb.save(OUTPUT_XLSX)
print(f"Excel report saved to: {OUTPUT_XLSX}")
