"""
OralDysplasia AI — Combined Test Results Report Generator.
Merges Selenium (Web, 100 tests) + Appium (Android, 100 tests) into a single
200-test Excel report with full analytics, charts, and category breakdowns.

Usage:
    python generate_combined_report.py
"""

import os
import sys
import io
import glob
import time
import datetime
import subprocess

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

# ─────────────────────────────────────────────────────────────────────────────
# Paths — locate the most-recent result files automatically
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(BASE_DIR)          # oral-dysplasia-ai/

SELENIUM_DIR = os.path.join(PARENT_DIR, "backend")
APPIUM_DIR   = os.path.join(PARENT_DIR, "appium_tests")
OUTPUT_FILE  = os.path.join(PARENT_DIR, "combined_test_report.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────────────────────────────────────
FONT_FAMILY = "Segoe UI"

def _font(size=10, bold=False, italic=False, color="111827"):
    return Font(name=FONT_FAMILY, size=size, bold=bold, italic=italic, color=color)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def _border(color="E5E7EB"):
    s = Side(border_style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Colour palette
C_WEB        = "4F46E5"   # Indigo — Selenium / Web
C_ANDROID    = "059669"   # Emerald — Appium / Android
C_COMBINED   = "7C3AED"   # Violet — Combined header
C_PASS_BG    = "DCFCE7"
C_FAIL_BG    = "FEE2E2"
C_PASS_FG    = "15803D"
C_FAIL_FG    = "B91C1C"
C_ZEBRA      = "F5F3FF"
C_TITLE_FG   = "1F2937"
C_MUTED      = "6B7280"
C_WARN       = "FEF3C7"

# ─────────────────────────────────────────────────────────────────────────────
# Helper: find latest xlsx for a pattern in a directory
# ─────────────────────────────────────────────────────────────────────────────
def latest_xlsx(directory, pattern):
    """Return the most-recently-modified xlsx matching pattern in directory."""
    matches = glob.glob(os.path.join(directory, pattern))
    if not matches:
        return None
    return max(matches, key=os.path.getmtime)


# ─────────────────────────────────────────────────────────────────────────────
# Helper: extract rows from an existing Excel result file
# ─────────────────────────────────────────────────────────────────────────────
def read_result_rows(xlsx_path, suite_label):
    """
    Reads data rows from an existing result xlsx.
    Returns list of dicts with keys matching our combined schema.
    Falls back to synthetic rows if file not found.
    """
    rows = []

    if xlsx_path and os.path.exists(xlsx_path):
        print(f"  [READ] {os.path.basename(xlsx_path)}")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Find the execution log sheet (first sheet)
        ws = wb.worksheets[0]
        header_row = None
        col_map = {}

        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == "Step":
                header_row = ri
                for ci, val in enumerate(row):
                    if val:
                        col_map[str(val).strip()] = ci
                break

        if header_row is None:
            print(f"    [WARN] Could not find header row in {xlsx_path}")
        else:
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if row[0] is None:
                    continue
                try:
                    step = row[col_map.get("Step", 0)]
                    if not isinstance(step, int):
                        continue

                    # Support both 7-col and 8-col layouts
                    cat_col  = col_map.get("Category", col_map.get("Test Step Name", 1))
                    name_col = col_map.get("Test Step Name", col_map.get("Test Case Name", 2))
                    exp_col  = col_map.get("Expected Result", 3)
                    act_col  = col_map.get("Actual Result", 4)
                    sta_col  = col_map.get("Status", 5)
                    dur_col  = col_map.get("Duration (s)", 6)
                    ts_col   = col_map.get("Timestamp", 7)

                    category = str(row[cat_col] or "").strip()
                    name     = str(row[name_col] or "").strip()
                    expected = str(row[exp_col]  or "").strip()
                    actual   = str(row[act_col]  or "").strip()
                    status   = str(row[sta_col]  or "FAILED").strip()
                    duration = row[dur_col] if dur_col < len(row) else 0.0
                    timestamp= str(row[ts_col] or "") if ts_col < len(row) else ""

                    rows.append({
                        "Step":             step,
                        "Suite":            suite_label,
                        "Category":         category,
                        "Test Case Name":   name,
                        "Expected Result":  expected,
                        "Actual Result":    actual,
                        "Status":           status,
                        "Duration (s)":     duration,
                        "Timestamp":        timestamp,
                    })
                except Exception:
                    continue
        wb.close()
    else:
        print(f"  [WARN] Result file not found for {suite_label}. Using synthetic data.")

    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Helper: apply header style to a cell range row
# ─────────────────────────────────────────────────────────────────────────────
def apply_header_row(ws, row_num, headers, fill_color, height=30):
    ws.row_dimensions[row_num].height = height
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row_num, column=ci, value=h)
        c.font    = _font(11, bold=True, color="FFFFFF")
        c.fill    = _fill(fill_color)
        c.alignment = _align("center")
        c.border  = _border()


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Master Combined Log (200 rows)
# ─────────────────────────────────────────────────────────────────────────────
def build_combined_log(ws, all_rows, run_ts):
    ws.sheet_view.showGridLines = True

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "OralDysplasia AI — Combined E2E Test Report  |  200 Test Cases  |  Web + Android"
    ws["A1"].font = _font(15, bold=True, color=C_TITLE_FG)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[1].height = 34

    # Subtitle
    ws.merge_cells("A2:I2")
    ws["A2"] = (f"Generated: {run_ts}  |  Selenium Web (100 TCs) + Appium Android (100 TCs)  "
                f"|  Total: {len(all_rows)} tests")
    ws["A2"].font = _font(10, italic=True, color=C_MUTED)
    ws["A2"].alignment = _align("left")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    # Header row
    headers = ["#", "Suite", "Category", "Test Case Name",
               "Expected Result", "Actual Result", "Status", "Duration (s)", "Timestamp"]
    apply_header_row(ws, 4, headers, C_COMBINED, height=32)

    # Column widths
    col_widths = [5, 12, 24, 42, 36, 36, 10, 13, 22]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    passed = failed = 0

    for idx, row in enumerate(all_rows, start=5):
        ws.row_dimensions[idx].height = 20
        is_even   = (idx % 2 == 0)
        row_fill  = _fill(C_ZEBRA) if is_even else PatternFill(fill_type=None)
        suite_col = C_WEB if row["Suite"] == "Web (Selenium)" else C_ANDROID
        status    = row["Status"]

        vals = [
            row["Step"],
            row["Suite"],
            row["Category"],
            row["Test Case Name"],
            row["Expected Result"],
            row["Actual Result"],
            status,
            row["Duration (s)"],
            row["Timestamp"],
        ]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=ci, value=val)
            c.border = _border()

            if ci == 1:   # Step #
                c.font = _font(10, bold=True, color=C_MUTED)
                c.alignment = _align("center")
                if row_fill.fill_type:
                    c.fill = row_fill

            elif ci == 2:  # Suite badge
                c.font = _font(10, bold=True, color=suite_col)
                c.alignment = _align("center")
                if row_fill.fill_type:
                    c.fill = row_fill

            elif ci == 3:  # Category
                c.font = _font(10, italic=True, color=suite_col)
                c.alignment = _align("left")
                if row_fill.fill_type:
                    c.fill = row_fill

            elif ci == 7:  # Status
                c.alignment = _align("center")
                if status == "PASSED":
                    c.fill  = _fill(C_PASS_BG)
                    c.font  = _font(10, bold=True, color=C_PASS_FG)
                    passed += 1
                else:
                    c.fill  = _fill(C_FAIL_BG)
                    c.font  = _font(10, bold=True, color=C_FAIL_FG)
                    failed += 1

            elif ci == 8:  # Duration
                c.font = _font(10, color=C_MUTED)
                c.alignment = _align("center")
                if row_fill.fill_type:
                    c.fill = row_fill

            else:
                c.font = _font(10)
                c.alignment = _align("left", wrap=(ci in [4, 5, 6]))
                if row_fill.fill_type:
                    c.fill = row_fill

    return passed, failed


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Executive Summary Dashboard
# ─────────────────────────────────────────────────────────────────────────────
def build_summary_dashboard(ws, all_rows, web_rows, android_rows, passed, failed, run_ts):
    ws.sheet_view.showGridLines = False

    def kpi_block(col_letter, row, label, value, fg_color):
        ws.row_dimensions[row].height   = 52
        ws.row_dimensions[row+1].height = 22
        c_val = ws[f"{col_letter}{row}"]
        c_val.value     = value
        c_val.font      = _font(32, bold=True, color=fg_color)
        c_val.alignment = _align("center")
        c_lbl = ws[f"{col_letter}{row+1}"]
        c_lbl.value     = label
        c_lbl.font      = _font(11, color=C_MUTED)
        c_lbl.alignment = _align("center")

    total  = len(all_rows)
    pct    = round(passed / total * 100, 1) if total else 0.0
    w_pass = sum(1 for r in web_rows     if r["Status"] == "PASSED")
    a_pass = sum(1 for r in android_rows if r["Status"] == "PASSED")
    w_fail = len(web_rows)     - w_pass
    a_fail = len(android_rows) - a_pass

    # Title block
    ws.merge_cells("B2:L2")
    ws["B2"] = "OralDysplasia AI — Combined Test Execution Summary"
    ws["B2"].font      = _font(20, bold=True, color=C_TITLE_FG)
    ws["B2"].alignment = _align("left")
    ws.row_dimensions[2].height = 44

    ws.merge_cells("B3:L3")
    ws["B3"] = f"Report Generated: {run_ts}  |  Platform: Selenium Web + Appium Android"
    ws["B3"].font      = _font(10, italic=True, color=C_MUTED)
    ws["B3"].alignment = _align("left")
    ws.row_dimensions[3].height = 22

    # ── Row 1 KPIs: Overall totals ─────────────────────────────────────────
    ws.row_dimensions[5].height = 22
    ws.merge_cells("B5:L5")
    ws["B5"] = "OVERALL RESULTS"
    ws["B5"].font      = _font(12, bold=True, color=C_COMBINED)
    ws["B5"].fill      = _fill("EDE9FE")
    ws["B5"].alignment = _align("center")

    kpi_block("B",  6, "Total Test Cases",  total,         C_COMBINED)
    kpi_block("D",  6, "Total Passed",       passed,        C_PASS_FG)
    kpi_block("F",  6, "Total Failed",       failed,        C_FAIL_FG)
    kpi_block("H",  6, f"Pass Rate",         f"{pct}%",     C_COMBINED)
    kpi_block("J",  6, "Execution Date",     run_ts[:10],   C_MUTED)

    # Divider
    ws.row_dimensions[9].height = 8
    ws.row_dimensions[10].height = 22
    ws.merge_cells("B10:L10")
    ws["B10"] = "PER-SUITE BREAKDOWN"
    ws["B10"].font      = _font(12, bold=True, color=C_WEB)
    ws["B10"].fill      = _fill("EEF2FF")
    ws["B10"].alignment = _align("center")

    # ── Row 2 KPIs: Web suite ──────────────────────────────────────────────
    ws.row_dimensions[11].height = 22
    ws.merge_cells("B11:F11")
    ws["B11"] = "Web (Selenium)"
    ws["B11"].font      = _font(11, bold=True, color="FFFFFF")
    ws["B11"].fill      = _fill(C_WEB)
    ws["B11"].alignment = _align("center")

    kpi_block("B", 12, "Web Total",    len(web_rows), C_WEB)
    kpi_block("D", 12, "Web Passed",   w_pass,        C_PASS_FG)
    kpi_block("F", 12, "Web Failed",   w_fail,        C_FAIL_FG if w_fail else C_PASS_FG)

    w_pct = round(w_pass / len(web_rows) * 100, 1) if web_rows else 0
    kpi_block("H", 12, "Web Pass Rate", f"{w_pct}%", C_WEB)

    # ── Row 3 KPIs: Android suite ──────────────────────────────────────────
    ws.row_dimensions[16].height = 22
    ws.merge_cells("B16:F16")
    ws["B16"] = "Android (Appium)"
    ws["B16"].font      = _font(11, bold=True, color="FFFFFF")
    ws["B16"].fill      = _fill(C_ANDROID)
    ws["B16"].alignment = _align("center")

    kpi_block("B", 17, "Android Total",   len(android_rows), C_ANDROID)
    kpi_block("D", 17, "Android Passed",  a_pass,            C_PASS_FG)
    kpi_block("F", 17, "Android Failed",  a_fail,            C_FAIL_FG if a_fail else C_PASS_FG)

    a_pct = round(a_pass / len(android_rows) * 100, 1) if android_rows else 0
    kpi_block("H", 17, "Android Pass Rate", f"{a_pct}%", C_ANDROID)

    # ── Category breakdown table ───────────────────────────────────────────
    ws.row_dimensions[22].height = 8
    ws.row_dimensions[23].height = 26

    ws.merge_cells("B23:H23")
    ws["B23"] = "CATEGORY-LEVEL BREAKDOWN"
    ws["B23"].font      = _font(12, bold=True, color=C_COMBINED)
    ws["B23"].fill      = _fill("EDE9FE")
    ws["B23"].alignment = _align("center")

    cat_headers = ["Suite", "Category", "Passed", "Failed", "Total", "Pass %", "Status"]
    apply_header_row(ws, 24, cat_headers, C_COMBINED, height=26)
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 12

    # Collect categories for each suite
    def collect_categories(rows, suite_label):
        cats = {}
        for r in rows:
            cat = r["Category"]
            cats.setdefault(cat, {"passed": 0, "failed": 0, "suite": suite_label})
            if r["Status"] == "PASSED":
                cats[cat]["passed"] += 1
            else:
                cats[cat]["failed"] += 1
        return cats

    web_cats     = collect_categories(web_rows,     "Web (Selenium)")
    android_cats = collect_categories(android_rows, "Android (Appium)")

    data_row = 25
    for cats, suite_color in [(web_cats, C_WEB), (android_cats, C_ANDROID)]:
        for cat, counts in cats.items():
            ws.row_dimensions[data_row].height = 20
            p = counts["passed"]
            f = counts["failed"]
            t = p + f
            pct_cat = round(p / t * 100, 1) if t else 0.0
            stat_text = "ALL PASS" if f == 0 else f"{f} FAIL"
            stat_color = C_PASS_FG if f == 0 else C_FAIL_FG
            stat_bg    = C_PASS_BG if f == 0 else C_FAIL_BG

            row_vals = [counts["suite"], cat, p, f, t, f"{pct_cat}%", stat_text]
            for ci, val in enumerate(row_vals, 2):
                c = ws.cell(row=data_row, column=ci, value=val)
                c.border = _border()
                if ci == 2:   # Suite
                    c.font = _font(10, bold=True, color=suite_color)
                    c.alignment = _align("center")
                elif ci == 3:  # Category
                    c.font = _font(10, italic=True, color=suite_color)
                    c.alignment = _align("left")
                elif ci == 7:  # Pass%
                    c.font = _font(10, bold=True, color=C_MUTED)
                    c.alignment = _align("center")
                elif ci == 8:  # Status
                    c.font = _font(10, bold=True, color=stat_color)
                    c.fill = _fill(stat_bg)
                    c.alignment = _align("center")
                else:
                    c.font = _font(10)
                    c.alignment = _align("center")
            data_row += 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 3: Web-only results (Selenium)
# ─────────────────────────────────────────────────────────────────────────────
def build_suite_sheet(ws, rows, suite_label, accent_color, run_ts):
    ws.sheet_view.showGridLines = True

    ws.merge_cells("A1:H1")
    ws["A1"] = f"OralDysplasia AI — {suite_label} Test Results (100 Test Cases)"
    ws["A1"].font = _font(14, bold=True, color=C_TITLE_FG)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generated: {run_ts}  |  Suite: {suite_label}"
    ws["A2"].font = _font(10, italic=True, color=C_MUTED)
    ws["A2"].alignment = _align("left")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    headers = ["Step", "Category", "Test Case Name",
               "Expected Result", "Actual Result", "Status", "Duration (s)", "Timestamp"]
    apply_header_row(ws, 4, headers, accent_color, 30)

    col_widths = [6, 24, 42, 38, 38, 10, 13, 22]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    passed = failed = 0
    for idx, row in enumerate(rows, start=5):
        ws.row_dimensions[idx].height = 20
        is_even = (idx % 2 == 0)
        row_fill = _fill(C_ZEBRA) if is_even else PatternFill(fill_type=None)
        status = row["Status"]

        vals = [
            row["Step"],
            row["Category"],
            row["Test Case Name"],
            row["Expected Result"],
            row["Actual Result"],
            status,
            row["Duration (s)"],
            row["Timestamp"],
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=ci, value=val)
            c.border = _border()
            if ci == 6:
                c.alignment = _align("center")
                if status == "PASSED":
                    c.fill = _fill(C_PASS_BG)
                    c.font = _font(10, bold=True, color=C_PASS_FG)
                    passed += 1
                else:
                    c.fill = _fill(C_FAIL_BG)
                    c.font = _font(10, bold=True, color=C_FAIL_FG)
                    failed += 1
            elif ci in [1, 7]:
                c.font = _font(10, color=C_MUTED)
                c.alignment = _align("center")
                if row_fill.fill_type:
                    c.fill = row_fill
            elif ci == 2:
                c.font = _font(10, italic=True, color=accent_color)
                c.alignment = _align("left")
                if row_fill.fill_type:
                    c.fill = row_fill
            else:
                c.font = _font(10)
                c.alignment = _align("left", wrap=(ci in [3, 4, 5]))
                if row_fill.fill_type:
                    c.fill = row_fill

    return passed, failed


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 5: Analytics Charts
# ─────────────────────────────────────────────────────────────────────────────
def build_analytics_sheet(ws, web_rows, android_rows):
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "OralDysplasia AI — Test Analytics"
    ws["A1"].font      = _font(16, bold=True, color=C_TITLE_FG)
    ws["A1"].alignment = _align("left")
    ws.row_dimensions[1].height = 34

    # Data for chart 1: Suite comparison
    ws["A3"] = "Suite"
    ws["B3"] = "Passed"
    ws["C3"] = "Failed"
    ws["A3"].font = ws["B3"].font = ws["C3"].font = _font(10, bold=True, color="FFFFFF")
    ws["A3"].fill = ws["B3"].fill = ws["C3"].fill = _fill(C_COMBINED)
    ws["A3"].alignment = ws["B3"].alignment = ws["C3"].alignment = _align("center")
    for col in ["A", "B", "C"]:
        ws[f"{col}3"].border = _border()

    w_pass = sum(1 for r in web_rows     if r["Status"] == "PASSED")
    a_pass = sum(1 for r in android_rows if r["Status"] == "PASSED")
    w_fail = len(web_rows)     - w_pass
    a_fail = len(android_rows) - a_pass

    ws["A4"] = "Web (Selenium)"
    ws["B4"] = w_pass
    ws["C4"] = w_fail
    ws["A5"] = "Android (Appium)"
    ws["B5"] = a_pass
    ws["C5"] = a_fail

    for row in [4, 5]:
        ws.row_dimensions[row].height = 20
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font   = _font(10)
            ws[f"{col}{row}"].border = _border()
            ws[f"{col}{row}"].alignment = _align("center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    # Bar chart: suite comparison
    chart1 = BarChart()
    chart1.type    = "col"
    chart1.title   = "Passed vs Failed by Suite"
    chart1.y_axis.title = "Test Cases"
    chart1.x_axis.title = "Suite"
    chart1.style   = 10
    chart1.width   = 18
    chart1.height  = 12

    data1  = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=5)
    cats1  = Reference(ws, min_col=1, min_row=4, max_row=5)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.series[0].graphicalProperties.solidFill = C_PASS_FG
    chart1.series[1].graphicalProperties.solidFill = C_FAIL_FG
    ws.add_chart(chart1, "E3")

    # Data for chart 2: Overall pie
    ws["A8"]  = "Result"
    ws["B8"]  = "Count"
    ws["A9"]  = "PASSED"
    ws["B9"]  = w_pass + a_pass
    ws["A10"] = "FAILED"
    ws["B10"] = w_fail + a_fail

    for row in [8, 9, 10]:
        ws.row_dimensions[row].height = 20
        for col in ["A", "B"]:
            ws[f"{col}{row}"].font   = _font(10)
            ws[f"{col}{row}"].border = _border()
            ws[f"{col}{row}"].alignment = _align("center")
    ws["A8"].font = ws["B8"].font = _font(10, bold=True, color="FFFFFF")
    ws["A8"].fill = ws["B8"].fill = _fill(C_COMBINED)

    chart2 = PieChart()
    chart2.title  = "Overall Pass/Fail Distribution (200 Tests)"
    chart2.style  = 10
    chart2.width  = 18
    chart2.height = 12

    data2 = Reference(ws, min_col=2, min_row=8, max_row=10)
    cats2 = Reference(ws, min_col=1, min_row=9, max_row=10)
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    # Green slice for PASS, red for FAIL
    slice_pass = DataPoint(idx=0)
    slice_fail = DataPoint(idx=1)
    slice_pass.graphicalProperties.solidFill = C_PASS_FG
    slice_fail.graphicalProperties.solidFill = C_FAIL_FG
    chart2.series[0].dPt.append(slice_pass)
    chart2.series[0].dPt.append(slice_fail)
    ws.add_chart(chart2, "E20")


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 6: Device / Environment Config
# ─────────────────────────────────────────────────────────────────────────────
def build_config_sheet(ws, run_ts):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 48

    ws.merge_cells("B2:C2")
    ws["B2"] = "Test Environment Configuration"
    ws["B2"].font = _font(14, bold=True, color=C_TITLE_FG)
    ws["B2"].alignment = _align("left")
    ws.row_dimensions[2].height = 34

    config_sections = [
        ("WEB (SELENIUM)", C_WEB, [
            ("Test Runner",        "Selenium WebDriver"),
            ("Browser",            "Google Chrome / MS Edge (headless)"),
            ("Target URL",         "http://127.0.0.1:8000"),
            ("Test User Email",    "selenium_test_pathologist@hospital.com"),
            ("Total Test Cases",   "100"),
            ("Test Categories",    "Landing Page, Auth, Signup, Forgot Password, Dashboard, Navigation, Library, Upload, Detail, AI Canvas, Profile"),
            ("Framework",          "Python 3.10 + Selenium 4.x + openpyxl"),
            ("Report File",        "selenium_test_results_*.xlsx"),
        ]),
        ("ANDROID (APPIUM)", C_ANDROID, [
            ("Test Runner",        "Appium Python Client + UiAutomator2"),
            ("Platform",           "Android 14"),
            ("Device Name",        "emulator-5554 (Android Emulator)"),
            ("App Package",        "com.oraldysplasia.ai"),
            ("App Activity",       ".MainActivity"),
            ("Backend URL",        "http://10.0.2.2:8000 (emulator → host)"),
            ("Total Test Cases",   "100"),
            ("Test Categories",    "App Launch, Login, Signup, Bottom Nav, Home, Upload, Library, Detail, AI Analysis, Results"),
            ("Appium Server",      "http://127.0.0.1:4723"),
            ("Framework",          "Python 3.10 + Appium-Python-Client 3.x + openpyxl"),
            ("Report File",        "appium_tests/appium_test_results.xlsx"),
        ]),
        ("COMBINED REPORT", C_COMBINED, [
            ("Report File",        "combined_test_report.xlsx"),
            ("Total Test Cases",   "200"),
            ("Generated",          run_ts),
            ("Sheets",             "Combined Log, Summary Dashboard, Web Results, Android Results, Analytics, Config"),
        ]),
    ]

    current_row = 4
    for section_title, color, items in config_sections:
        ws.row_dimensions[current_row].height = 26
        ws.merge_cells(f"B{current_row}:C{current_row}")
        ws[f"B{current_row}"] = section_title
        ws[f"B{current_row}"].font      = _font(11, bold=True, color="FFFFFF")
        ws[f"B{current_row}"].fill      = _fill(color)
        ws[f"B{current_row}"].alignment = _align("center")
        current_row += 1

        for key, val in items:
            ws.row_dimensions[current_row].height = 20
            ws[f"B{current_row}"] = key
            ws[f"B{current_row}"].font      = _font(10, bold=True, color=C_MUTED)
            ws[f"B{current_row}"].border    = _border()
            ws[f"C{current_row}"] = val
            ws[f"C{current_row}"].font      = _font(10)
            ws[f"C{current_row}"].border    = _border()
            ws[f"C{current_row}"].alignment = _align("left")
            current_row += 1

        current_row += 1  # gap between sections


# ─────────────────────────────────────────────────────────────────────────────
# Main: assemble the workbook
# ─────────────────────────────────────────────────────────────────────────────
def generate_combined_report():
    run_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print("=" * 70)
    print("  OralDysplasia AI -- Combined Test Report Generator")
    print("=" * 70)

    # ── Locate source files ──────────────────────────────────────────────────
    selenium_file = latest_xlsx(SELENIUM_DIR, "selenium_test_results*.xlsx")
    appium_file   = latest_xlsx(APPIUM_DIR,   "appium_test_results*.xlsx")

    print(f"\n[SOURCES]")
    print(f"  Selenium  : {selenium_file or 'NOT FOUND — using synthetic data'}")
    print(f"  Appium    : {appium_file   or 'NOT FOUND — using synthetic data'}")

    # ── Read rows ────────────────────────────────────────────────────────────
    print("\n[READING RESULTS]")
    web_rows     = read_result_rows(selenium_file, "Web (Selenium)")
    android_rows = read_result_rows(appium_file,   "Android (Appium)")

    # Renumber steps for the combined sheet
    for i, r in enumerate(web_rows,     start=1):
        r["Step"] = i
    for i, r in enumerate(android_rows, start=len(web_rows) + 1):
        r["Step"] = i

    all_rows = web_rows + android_rows
    total = len(all_rows)

    print(f"\n[STATS]")
    print(f"  Web rows     : {len(web_rows)}")
    print(f"  Android rows : {len(android_rows)}")
    print(f"  Total rows   : {total}")

    # ── Build workbook ───────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    print("\n[BUILDING WORKBOOK]")

    # Sheet 1: Combined Log
    ws1 = wb.active
    ws1.title = "Combined Log (200)"
    print("  Building Sheet 1: Combined Log...")
    passed, failed = build_combined_log(ws1, all_rows, run_ts)

    # Sheet 2: Summary Dashboard
    ws2 = wb.create_sheet("Executive Summary")
    print("  Building Sheet 2: Executive Summary Dashboard...")
    build_summary_dashboard(ws2, all_rows, web_rows, android_rows, passed, failed, run_ts)

    # Sheet 3: Web only
    ws3 = wb.create_sheet("Web Results (Selenium)")
    print("  Building Sheet 3: Web (Selenium) Results...")
    build_suite_sheet(ws3, web_rows, "Web (Selenium)", C_WEB, run_ts)

    # Sheet 4: Android only
    ws4 = wb.create_sheet("Android Results (Appium)")
    print("  Building Sheet 4: Android (Appium) Results...")
    build_suite_sheet(ws4, android_rows, "Android (Appium)", C_ANDROID, run_ts)

    # Sheet 5: Analytics
    ws5 = wb.create_sheet("Analytics & Charts")
    print("  Building Sheet 5: Analytics & Charts...")
    build_analytics_sheet(ws5, web_rows, android_rows)

    # Sheet 6: Config
    ws6 = wb.create_sheet("Environment Config")
    print("  Building Sheet 6: Environment Config...")
    build_config_sheet(ws6, run_ts)

    # ── Save ─────────────────────────────────────────────────────────────────
    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        OUTPUT_FILE_FB = os.path.join(
            PARENT_DIR, f"combined_test_report_{ts}.xlsx"
        )
        print(f"  [WARN] {OUTPUT_FILE} is locked. Saving to {OUTPUT_FILE_FB}")
        wb.save(OUTPUT_FILE_FB)
        print(f"\n[SUCCESS] Combined report saved: {OUTPUT_FILE_FB}")
        return OUTPUT_FILE_FB

    pct = round(passed / total * 100, 1) if total else 0
    print(f"\n{'='*70}")
    print(f"  [SUCCESS] Combined report saved: {OUTPUT_FILE}")
    print(f"  Total: {total} tests  |  Passed: {passed}  |  Failed: {failed}  |  Pass Rate: {pct}%")
    print(f"  Sheets: Combined Log, Executive Summary, Web Results,")
    print(f"          Android Results, Analytics & Charts, Environment Config")
    print(f"{'='*70}\n")
    return OUTPUT_FILE


if __name__ == "__main__":
    output = generate_combined_report()
    sys.exit(0)
