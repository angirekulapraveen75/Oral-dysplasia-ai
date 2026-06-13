import os
import openpyxl
from fpdf import FPDF
from fpdf.fonts import FontFace

def clean_text(txt):
    if not txt:
        return ""
    txt = str(txt)
    replacements = {
        "\u2013": "-",
        "\u2014": "-",
        "\u2018": "'",
        "\u2019": "'",
        "\u201c": '"',
        "\u201d": '"',
        "\u2022": "*",
    }
    for orig, repl in replacements.items():
        txt = txt.replace(orig, repl)
    return txt.encode("latin-1", errors="ignore").decode("latin-1")

def convert_excel_to_pdf(excel_path="selenium_test_results.xlsx", pdf_path="selenium_test_results.pdf"):
    print(f"Loading Excel file: {excel_path}...")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file '{excel_path}' not found. Please run the Selenium tests first.")
        
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Extract Title and Subtitle
    title = clean_text(ws["A1"].value or "OralDysplasia AI — End-to-End System Validation Report")
    subtitle = clean_text(ws["A2"].value or "System Validation Report")
    
    # Read Headers and Data rows
    headers = []
    data_rows = []
    
    # Headers are in Row 4
    for col_idx in range(1, 8):
        headers.append(clean_text(ws.cell(row=4, column=col_idx).value))
        
    # Data is in Row 5 onwards
    row_idx = 5
    while True:
        step = ws.cell(row=row_idx, column=1).value
        if step is None:
            break
        row_data = []
        for col_idx in range(1, 8):
            row_data.append(clean_text(ws.cell(row=row_idx, column=col_idx).value))
        data_rows.append(row_data)
        row_idx += 1
        
    print(f"Read {len(data_rows)} test steps from Excel.")
    
    # Generate PDF in Landscape format (A4)
    pdf = FPDF(orientation="landscape", unit="mm", format="A4")
    pdf.set_margins(15, 15, 15)
    pdf.add_page()
    
    # Header Banner - Title & Subtitle
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(31, 41, 55) # Dark gray #1F2937
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align="L")
    
    pdf.set_font("Helvetica", "I", 10)
    pdf.set_text_color(107, 114, 128) # Gray #6B7280
    pdf.cell(0, 8, subtitle, new_x="LMARGIN", new_y="NEXT", align="L")
    
    pdf.ln(5)
    
    # Column widths summing to 267 (A4 landscape usable width)
    col_widths = (12, 55, 60, 75, 22, 20, 23)
    
    # Use fpdf2's built-in table layout with beautiful styles
    with pdf.table(
        col_widths=col_widths,
        text_align=("CENTER", "LEFT", "LEFT", "LEFT", "CENTER", "CENTER", "CENTER"),
        line_height=7,
        padding=3
    ) as table:
        # Header Row
        header_row = table.row()
        pdf.set_font("Helvetica", "B", 9)
        header_fill = FontFace(color=(255, 255, 255), fill_color=(79, 70, 229)) # Indigo #4F46E5
        
        for header in headers:
            header_row.cell(header, style=header_fill)
            
        # Data Rows
        pdf.set_font("Helvetica", "", 8)
        
        for idx, r in enumerate(data_rows):
            data_row = table.row()
            # Zebra striping
            bg_color = (249, 250, 251) if idx % 2 == 1 else (255, 255, 255)
            
            for col_idx, val in enumerate(r):
                # Color code Status column
                if col_idx == 4:
                    if val == "PASSED":
                        # Light green bg, dark green text
                        status_style = FontFace(color=(21, 128, 61), fill_color=(220, 252, 231))
                        data_row.cell(val, style=status_style)
                    else:
                        # Light red bg, dark red text
                        status_style = FontFace(color=(185, 28, 28), fill_color=(254, 226, 226))
                        data_row.cell(val, style=status_style)
                else:
                    style = FontFace(color=(0, 0, 0), fill_color=bg_color)
                    data_row.cell(val, style=style)
                    
    pdf.output(pdf_path)
    print(f"PDF saved successfully as '{pdf_path}'.")

if __name__ == "__main__":
    convert_excel_to_pdf()
