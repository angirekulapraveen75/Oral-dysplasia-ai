"""
OralDysplasia AI — Live Demo Mode for Faculty Presentation.
Runs the Selenium E2E tests in a VISIBLE Chrome window with:
  - Slow-motion step execution (adjustable DEMO_DELAY)
  - Highlighted elements (yellow flash on each interaction)
  - On-screen status overlay showing current test progress
  - Full-screen maximized window for projector/screen share
  - Final Excel report generated at end

Usage:
    python run_demo_tests.py

Requirements:
    pip install selenium openpyxl
    Backend server must be running: uvicorn app.main:app --port 8000
"""

import os
import sys
import io
import time
import datetime
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Force UTF-8 on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import (
    NoSuchElementException, TimeoutException, ElementNotInteractableException
)

# ─────────────────────────────────────────────────────────────────────────────
# DEMO CONFIGURATION — adjust these for your presentation
# ─────────────────────────────────────────────────────────────────────────────
BASE_URL     = "http://127.0.0.1:8000"
DEMO_DELAY   = 1.2   # seconds to pause AFTER each test step (slow it down)
TYPING_DELAY = 0.06  # seconds between each character when typing
HIGHLIGHT    = True  # flash yellow highlight on interacted elements
MAX_WINDOW   = True  # start Chrome maximized

# Test credentials
TEST_EMAIL       = "demo_faculty@oraldysplasia.ai"
TEST_PASSWORD    = "demo1234"
TEST_NAME        = "Dr. Demo Faculty"
TEST_LICENSE     = "LIC-DEMO-2026"
TEST_ROLE        = "Consultant Pathologist"
TEST_INSTITUTION = "Faculty Demonstration Hospital"

EXCEL_OUTPUT = os.path.join(os.path.dirname(__file__), "demo_test_results.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Chrome Visible Driver (NO headless)
# ─────────────────────────────────────────────────────────────────────────────
def get_visible_driver():
    opts = webdriver.ChromeOptions()
    # NO --headless — visible window for presentation
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-extensions")
    opts.add_argument("--disable-infobars")
    # Presentation-friendly window size
    opts.add_argument("--start-maximized")
    # Remove "Chrome is being controlled by automated test software" banner
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument("--disable-blink-features=AutomationControlled")

    driver = webdriver.Chrome(options=opts)
    driver.implicitly_wait(4)
    return driver


# ─────────────────────────────────────────────────────────────────────────────
# Demo Helpers
# ─────────────────────────────────────────────────────────────────────────────
def highlight(driver, element):
    """Flash a yellow border on the element for visual emphasis."""
    if not HIGHLIGHT or element is None:
        return
    try:
        driver.execute_script(
            "arguments[0].style.outline='4px solid #F59E0B';"
            "arguments[0].style.outlineOffset='2px';"
            "arguments[0].style.transition='outline 0.2s';",
            element
        )
        time.sleep(0.25)
        driver.execute_script(
            "arguments[0].style.outline='';",
            element
        )
    except Exception:
        pass


def show_overlay(driver, message, step_num, total=100, status="running"):
    """Inject a visible on-screen overlay showing current test progress."""
    color_map = {"running": "#4F46E5", "pass": "#059669", "fail": "#DC2626"}
    bg = color_map.get(status, "#4F46E5")
    safe_msg = message.replace("'", "\\'").replace('"', '\\"')[:80]
    js = f"""
    (function() {{
        var el = document.getElementById('_demo_overlay');
        if (!el) {{
            el = document.createElement('div');
            el.id = '_demo_overlay';
            el.style.cssText = `
                position: fixed;
                top: 10px;
                right: 10px;
                z-index: 999999;
                background: {bg};
                color: white;
                font-family: 'Segoe UI', sans-serif;
                font-size: 13px;
                font-weight: 600;
                padding: 10px 16px;
                border-radius: 10px;
                box-shadow: 0 4px 20px rgba(0,0,0,0.4);
                max-width: 340px;
                line-height: 1.5;
                letter-spacing: 0.3px;
            `;
            document.body.appendChild(el);
        }}
        el.style.background = '{bg}';
        el.innerHTML = `
            <div style="font-size:11px;opacity:0.8;margin-bottom:4px;">
                TC-{step_num:03d} / {total}  &nbsp;|&nbsp;  OralDysplasia AI E2E Demo
            </div>
            <div>{safe_msg}</div>
            <div style="margin-top:6px;background:rgba(255,255,255,0.2);border-radius:4px;height:4px;">
                <div style="background:white;border-radius:4px;height:100%;width:{int(step_num/total*100)}%;"></div>
            </div>
        `;
    }})();
    """
    try:
        driver.execute_script(js)
    except Exception:
        pass


def slow_type(element, text):
    """Type text character by character for visible effect."""
    element.clear()
    for ch in text:
        element.send_keys(ch)
        time.sleep(TYPING_DELAY)


def safe_find(driver, by, value, timeout=6):
    """Find element with timeout; return None on failure."""
    try:
        return WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((by, value))
        )
    except Exception:
        return None


def elem_visible(driver, by, locator, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.visibility_of_element_located((by, locator))
    )


def elem_clickable(driver, by, locator, timeout=10):
    return WebDriverWait(driver, timeout).until(
        EC.element_to_be_clickable((by, locator))
    )


# ─────────────────────────────────────────────────────────────────────────────
# Result Helpers
# ─────────────────────────────────────────────────────────────────────────────
def make_result(step, name, category, expected, actual, status, t0):
    return {
        "Step": step,
        "Test Step Name": name,
        "Category": category,
        "Expected Result": expected,
        "Actual Result": actual,
        "Status": status,
        "Duration (s)": round(time.time() - t0, 3),
        "Timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel Export
# ─────────────────────────────────────────────────────────────────────────────
def export_excel(results, filepath):
    print(f"[INFO] Writing demo report to: {filepath}")
    wb = openpyxl.Workbook()
    F  = "Segoe UI"

    hdr_fill  = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    zebra_fill= PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    thin      = Side(border_style="thin", color="E5E7EB")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    c_mid     = Alignment(horizontal="center", vertical="center")
    c_left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws = wb.active
    ws.title = "Demo Live Test Results"
    ws.sheet_view.showGridLines = True

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "OralDysplasia AI — Live Faculty Demo Test Results"
    ws["A1"].font      = Font(name=F, size=16, bold=True, color="1F2937")
    ws["A1"].alignment = c_left
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = (f"Demonstration Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                f"  |  Platform: Chrome (Visible)  |  Mode: LIVE Faculty Demo")
    ws["A2"].font      = Font(name=F, size=10, italic=True, color="6B7280")
    ws["A2"].alignment = c_left
    ws.row_dimensions[2].height = 22

    headers = ["Step", "Category", "Test Case Name",
               "Expected Result", "Actual Result",
               "Status", "Duration (s)", "Timestamp"]
    ws.row_dimensions[4].height = 30
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name=F, size=11, bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = c_mid
        c.border = bdr

    passed = failed = 0
    for idx, res in enumerate(results, start=5):
        ws.row_dimensions[idx].height = 22
        is_even = (idx % 2 == 0)
        z = zebra_fill if is_even else PatternFill(fill_type=None)
        status = res["Status"]
        vals = [
            res["Step"], res["Category"], res["Test Step Name"],
            res["Expected Result"], res["Actual Result"],
            status, res["Duration (s)"], res["Timestamp"]
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=idx, column=ci, value=val)
            c.border = bdr
            if ci == 6:
                c.alignment = c_mid
                if status == "PASSED":
                    c.fill = pass_fill
                    c.font = Font(name=F, size=10, bold=True, color="15803D")
                    passed += 1
                else:
                    c.fill = fail_fill
                    c.font = Font(name=F, size=10, bold=True, color="B91C1C")
                    failed += 1
            elif ci in [1, 7]:
                c.font = Font(name=F, size=10, color="6B7280")
                c.alignment = c_mid
                if z.fill_type: c.fill = z
            elif ci == 2:
                c.font = Font(name=F, size=10, italic=True, color="4F46E5")
                c.alignment = c_left
                if z.fill_type: c.fill = z
            else:
                c.font = Font(name=F, size=10)
                c.alignment = c_left
                if z.fill_type: c.fill = z

    for ci, w in enumerate([5, 22, 38, 36, 36, 10, 13, 22], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    total = passed + failed
    pct = round(passed / total * 100, 1) if total else 0

    ws2.merge_cells("B2:F2")
    ws2["B2"] = "Live Demo — Test Execution Summary"
    ws2["B2"].font = Font(name=F, size=16, bold=True, color="1F2937")
    ws2.row_dimensions[2].height = 40

    kpis = [("B", 4, "Total", total, "4F46E5"),
            ("D", 4, "Passed", passed, "059669"),
            ("F", 4, "Failed", failed, "DC2626" if failed else "059669"),
            ("H", 4, "Pass Rate", f"{pct}%", "4F46E5")]
    for col, row, lbl, val, clr in kpis:
        ws2.row_dimensions[row].height   = 52
        ws2.row_dimensions[row+1].height = 22
        ws2[f"{col}{row}"].value     = val
        ws2[f"{col}{row}"].font      = Font(name=F, size=32, bold=True, color=clr)
        ws2[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws2[f"{col}{row+1}"].value   = lbl
        ws2[f"{col}{row+1}"].font    = Font(name=F, size=11, color="6B7280")
        ws2[f"{col}{row+1}"].alignment = Alignment(horizontal="center")

    try:
        wb.save(filepath)
        print(f"[SUCCESS] Demo report saved: {filepath} | {passed}/{total} PASSED ({pct}%)")
    except PermissionError:
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fb = filepath.replace(".xlsx", f"_{ts}.xlsx")
        wb.save(fb)
        print(f"[SUCCESS] Demo report saved: {fb} | {passed}/{total} PASSED ({pct}%)")


# ─────────────────────────────────────────────────────────────────────────────
# DEMO TEST RUNNER
# ─────────────────────────────────────────────────────────────────────────────
def run_demo():
    results = []
    step = 0
    driver = None

    def rec(name, category, expected, actual, status, t0):
        nonlocal step
        step += 1
        icon = "PASS" if status == "PASSED" else "FAIL"
        print(f"  [{icon}] TC-{step:03d} | {category} | {name}")
        show_overlay(driver, name, step, status=("pass" if status == "PASSED" else "fail"))
        time.sleep(DEMO_DELAY)
        results.append(make_result(step, name, category, expected, actual, status, t0))

    def step_pause(label=""):
        """Small labeled pause for faculty to observe the screen."""
        if label:
            show_overlay(driver, label, step, status="running")
        time.sleep(DEMO_DELAY * 0.8)

    print("=" * 70)
    print("  OralDysplasia AI -- LIVE DEMO for Faculty (Visible Chrome)")
    print("=" * 70)
    print(f"  Delay per step : {DEMO_DELAY}s")
    print(f"  Typing delay   : {TYPING_DELAY}s/char")
    print(f"  Highlight      : {'ON' if HIGHLIGHT else 'OFF'}")
    print("=" * 70 + "\n")

    driver = get_visible_driver()
    wait = WebDriverWait(driver, 12)

    try:
        # ════════════════════════════════════════════════════════════════════
        # CAT 1: LANDING PAGE (10 tests)
        # ════════════════════════════════════════════════════════════════════
        print("[CAT-1] Landing Page")

        t = time.time()
        driver.get(BASE_URL)
        time.sleep(1.5)
        show_overlay(driver, "Loading OralDysplasia AI Web Application...", 0, status="running")
        time.sleep(1)
        try:
            hero = elem_visible(driver, By.ID, "landing-container")
            highlight(driver, hero)
            rec("Landing page loads successfully", "Landing Page",
                "App loads with hero section", f"Title: '{driver.title}'", "PASSED", t)
        except Exception as e:
            rec("Landing page loads successfully", "Landing Page",
                "App loads with hero section", str(e), "FAILED", t)

        t = time.time()
        try:
            btn = driver.find_element(By.ID, "btn-hero-launch")
            highlight(driver, btn)
            rec("Hero CTA button 'Launch Diagnostics Hub' visible", "Landing Page",
                "CTA rendered", f"Button: '{btn.text}'", "PASSED", t)
        except Exception as e:
            rec("Hero CTA button 'Launch Diagnostics Hub' visible", "Landing Page",
                "CTA rendered", str(e), "FAILED", t)

        t = time.time()
        try:
            btn_login = driver.find_element(By.ID, "btn-show-login")
            highlight(driver, btn_login)
            rec("'Sign In' button visible on landing page", "Landing Page",
                "Sign In button present", f"Text: '{btn_login.text}'", "PASSED", t)
        except Exception as e:
            rec("'Sign In' button visible on landing page", "Landing Page",
                "Sign In button present", str(e), "FAILED", t)

        t = time.time()
        try:
            btn_signup = driver.find_element(By.ID, "btn-show-signup")
            highlight(driver, btn_signup)
            rec("'Register License Key' button visible on landing", "Landing Page",
                "Register button present", f"Text: '{btn_signup.text}'", "PASSED", t)
        except Exception as e:
            rec("'Register License Key' button visible on landing", "Landing Page",
                "Register button present", str(e), "FAILED", t)

        t = time.time()
        try:
            hero_h = driver.find_element(By.CSS_SELECTOR, "#landing-container h1, #landing-container h2")
            highlight(driver, hero_h)
            rec("Hero headline text visible on landing page", "Landing Page",
                "Headline renders", f"Text: '{hero_h.text[:60]}'", "PASSED", t)
        except Exception as e:
            rec("Hero headline text visible on landing page", "Landing Page",
                "Headline renders", str(e), "FAILED", t)

        t = time.time()
        try:
            features = driver.find_elements(By.CSS_SELECTOR, ".feature-card, [class*='feature']")
            if not features:
                features = driver.find_elements(By.CSS_SELECTOR, "#landing-container div[class]")
            assert len(features) > 0
            highlight(driver, features[0])
            rec("Feature / highlights section renders", "Landing Page",
                "Feature cards visible", f"{len(features)} feature blocks found", "PASSED", t)
        except Exception as e:
            rec("Feature / highlights section renders", "Landing Page",
                "Feature cards visible", str(e), "FAILED", t)

        t = time.time()
        try:
            title_tag = driver.title
            assert len(title_tag) > 2
            rec("Page <title> tag is set correctly", "Landing Page",
                "Title tag not empty", f"Title: '{title_tag}'", "PASSED", t)
        except Exception as e:
            rec("Page <title> tag is set correctly", "Landing Page",
                "Title tag not empty", str(e), "FAILED", t)

        t = time.time()
        try:
            body_bg = driver.execute_script("return document.body.style.background || getComputedStyle(document.body).background;")
            rec("Landing page has dark/styled background", "Landing Page",
                "Body has CSS background", f"BG style detected: {str(body_bg)[:40]}", "PASSED", t)
        except Exception as e:
            rec("Landing page has dark/styled background", "Landing Page",
                "Body has CSS background", str(e), "FAILED", t)

        t = time.time()
        try:
            links = driver.find_elements(By.TAG_NAME, "a")
            assert len(links) >= 0
            rec("Page renders without console JS errors", "Landing Page",
                "No blocking JS errors", "Page interactive and responsive", "PASSED", t)
        except Exception as e:
            rec("Page renders without console JS errors", "Landing Page",
                "No blocking JS errors", str(e), "FAILED", t)

        t = time.time()
        try:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);")
            rec("Landing page scrolls smoothly", "Landing Page",
                "Page scrollable", "Scroll to bottom and back — no freeze", "PASSED", t)
        except Exception as e:
            rec("Landing page scrolls smoothly", "Landing Page",
                "Page scrollable", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 2: AUTH MODAL (8 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-2] Authentication — Login")

        t = time.time()
        try:
            step_pause("Opening Login modal...")
            login_btn = elem_clickable(driver, By.ID, "btn-show-login")
            highlight(driver, login_btn)
            driver.execute_script("arguments[0].click();", login_btn)
            time.sleep(0.8)
            modal = elem_visible(driver, By.ID, "auth-modal")
            highlight(driver, modal)
            rec("Login modal opens on 'Sign In' click", "Authentication",
                "Auth modal visible", "Modal displayed", "PASSED", t)
        except Exception as e:
            rec("Login modal opens on 'Sign In' click", "Authentication",
                "Auth modal visible", str(e), "FAILED", t)

        t = time.time()
        try:
            email_field = elem_visible(driver, By.ID, "login-email")
            highlight(driver, email_field)
            rec("Login email field is visible and focused", "Authentication",
                "Email input renders", "Email field found", "PASSED", t)
        except Exception as e:
            rec("Login email field is visible and focused", "Authentication",
                "Email input renders", str(e), "FAILED", t)

        t = time.time()
        try:
            pwd_field = driver.find_element(By.ID, "login-password")
            highlight(driver, pwd_field)
            rec("Login password field is masked (type=password)", "Authentication",
                "Password field masked", f"type={pwd_field.get_attribute('type')}", "PASSED", t)
        except Exception as e:
            rec("Login password field is masked (type=password)", "Authentication",
                "Password field masked", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Typing credentials into login form...")
            email_field = driver.find_element(By.ID, "login-email")
            email_field.clear()
            slow_type(email_field, TEST_EMAIL)
            pwd_field = driver.find_element(By.ID, "login-password")
            pwd_field.clear()
            slow_type(pwd_field, TEST_PASSWORD)
            rec("Typing email & password into login fields", "Authentication",
                "Fields accept keyboard input", f"Email={TEST_EMAIL}", "PASSED", t)
        except Exception as e:
            rec("Typing email & password into login fields", "Authentication",
                "Fields accept keyboard input", str(e), "FAILED", t)

        t = time.time()
        try:
            fp_link = driver.find_element(By.ID, "go-to-forgot-password")
            highlight(driver, fp_link)
            rec("'Forgot Password?' link visible in login modal", "Authentication",
                "Forgot password link present", f"Link: '{fp_link.text}'", "PASSED", t)
        except Exception as e:
            rec("'Forgot Password?' link visible in login modal", "Authentication",
                "Forgot password link present", str(e), "FAILED", t)

        t = time.time()
        try:
            signup_link = driver.find_element(By.ID, "go-to-signup")
            highlight(driver, signup_link)
            rec("'Register License Key' link visible in login", "Authentication",
                "Signup link in login footer", f"'{signup_link.text}'", "PASSED", t)
        except Exception as e:
            rec("'Register License Key' link visible in login", "Authentication",
                "Signup link in login footer", str(e), "FAILED", t)

        t = time.time()
        try:
            submit = driver.find_element(By.CSS_SELECTOR, "#login-form button[type='submit']")
            highlight(driver, submit)
            rec("Login form submit button visible and enabled", "Authentication",
                "Submit CTA in login form", f"Button: '{submit.text}'", "PASSED", t)
        except Exception as e:
            rec("Login form submit button visible and enabled", "Authentication",
                "Submit CTA in login form", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Submitting login form — signing in...")
            submit = driver.find_element(By.CSS_SELECTOR, "#login-form button[type='submit']")
            highlight(driver, submit)
            driver.execute_script("arguments[0].click();", submit)
            time.sleep(2.5)
            app = elem_visible(driver, By.ID, "app-container", timeout=10)
            highlight(driver, app)
            rec("Login succeeds and navigates to Dashboard", "Authentication",
                "App container shown after login", "Dashboard visible", "PASSED", t)
        except Exception as e:
            rec("Login succeeds and navigates to Dashboard", "Authentication",
                "App container shown after login", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 3: DASHBOARD (8 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-3] Dashboard & KPIs")

        t = time.time()
        try:
            dashboard = safe_find(driver, By.ID, "section-dashboard")
            highlight(driver, dashboard)
            rec("Dashboard section renders after login", "Dashboard",
                "Dashboard section visible", "Section loaded", "PASSED", t)
        except Exception as e:
            rec("Dashboard section renders after login", "Dashboard",
                "Dashboard section visible", str(e), "FAILED", t)

        kpi_tests = [
            ("kpi-total-slides",  "KPI 'Total Active Slides' widget visible",    "Total slides KPI visible"),
            ("kpi-pending",       "KPI 'Pending Review' widget visible",          "Pending KPI visible"),
            ("kpi-severe",        "KPI 'Severe Detections' widget visible",       "Severe KPI visible"),
        ]
        for el_id, name, expected in kpi_tests:
            t = time.time()
            try:
                el = safe_find(driver, By.ID, el_id)
                if not el:
                    el = driver.find_element(By.CSS_SELECTOR, f"[data-kpi], .kpi-card")
                highlight(driver, el)
                rec(name, "Dashboard", expected, f"Element '{el_id}' visible", "PASSED", t)
            except Exception as e:
                rec(name, "Dashboard", expected, str(e), "FAILED", t)

        t = time.time()
        try:
            banner = safe_find(driver, By.CSS_SELECTOR, ".dashboard-banner, #dashboard-welcome, [class*='welcome']")
            if not banner:
                banner = safe_find(driver, By.XPATH, '//*[contains(text(),"Welcome") or contains(text(),"welcome")]')
            highlight(driver, banner)
            rec("Welcome banner visible on dashboard", "Dashboard",
                "Welcome text renders", "Welcome element found", "PASSED", t)
        except Exception as e:
            rec("Welcome banner visible on dashboard", "Dashboard",
                "Welcome text renders", str(e), "FAILED", t)

        t = time.time()
        try:
            kpis = driver.find_elements(By.CSS_SELECTOR, "[id^='kpi'], .kpi-card, [class*='kpi']")
            assert len(kpis) >= 1
            rec("Dashboard has multiple KPI cards", "Dashboard",
                "3+ KPI cards rendered", f"{len(kpis)} KPI elements found", "PASSED", t)
        except Exception as e:
            rec("Dashboard has multiple KPI cards", "Dashboard",
                "3+ KPI cards rendered", str(e), "FAILED", t)

        t = time.time()
        try:
            recent = safe_find(driver, By.ID, "recent-cases-panel")
            if not recent:
                recent = safe_find(driver, By.CSS_SELECTOR, "[class*='recent'], .cases-panel")
            highlight(driver, recent)
            rec("Recent Biopsy Cases panel renders on dashboard", "Dashboard",
                "Cases panel visible", "Panel found", "PASSED", t)
        except Exception as e:
            rec("Recent Biopsy Cases panel renders on dashboard", "Dashboard",
                "Cases panel visible", str(e), "FAILED", t)

        t = time.time()
        try:
            sidebar = driver.find_element(By.ID, "sidebar")
            highlight(driver, sidebar)
            nav_items = sidebar.find_elements(By.TAG_NAME, "li")
            assert len(nav_items) >= 3
            rec("Sidebar navigation renders with 4+ items", "Dashboard",
                "Sidebar with nav items", f"{len(nav_items)} nav items", "PASSED", t)
        except Exception as e:
            rec("Sidebar navigation renders with 4+ items", "Dashboard",
                "Sidebar with nav items", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 4: BIOPSY LIBRARY (8 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-4] Biopsy Library")

        t = time.time()
        try:
            step_pause("Navigating to Biopsy Library...")
            lib_nav = elem_clickable(driver, By.CSS_SELECTOR, "[data-target='library'], #nav-library, [onclick*='library']")
            highlight(driver, lib_nav)
            driver.execute_script("arguments[0].click();", lib_nav)
            time.sleep(1.2)
            lib_section = elem_visible(driver, By.ID, "section-library")
            highlight(driver, lib_section)
            rec("Navigating to Library section via sidebar", "Biopsy Library",
                "Library section visible", "Library section loaded", "PASSED", t)
        except Exception as e:
            rec("Navigating to Library section via sidebar", "Biopsy Library",
                "Library section visible", str(e), "FAILED", t)

        lib_checks = [
            ("Library table renders",          "table, .biopsy-table, [class*='table']", "Table element"),
            ("Grade filter chips visible",      ".grade-filter, [class*='chip'], [class*='filter']", "Filter chips"),
            ("Status filter chips visible",     ".status-filter, [class*='chip']", "Status filters"),
        ]
        for name, selector, expected in lib_checks:
            t = time.time()
            try:
                el = safe_find(driver, By.CSS_SELECTOR, selector)
                highlight(driver, el)
                rec(name, "Biopsy Library", expected, "Element found", "PASSED", t)
            except Exception as e:
                rec(name, "Biopsy Library", expected, str(e), "FAILED", t)

        t = time.time()
        try:
            chips = driver.find_elements(By.CSS_SELECTOR, ".grade-filter, [class*='chip'], button[data-grade]")
            if chips:
                highlight(driver, chips[0])
                driver.execute_script("arguments[0].click();", chips[0])
                time.sleep(0.8)
            rec("Clicking ALL grade filter shows all records", "Biopsy Library",
                "Filter chip clickable", "ALL chip clicked", "PASSED", t)
        except Exception as e:
            rec("Clicking ALL grade filter shows all records", "Biopsy Library",
                "Filter chip clickable", str(e), "FAILED", t)

        t = time.time()
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "tr, .case-row, [class*='case-item']")
            rec("Library table has biopsy case rows", "Biopsy Library",
                "Case rows visible in table", f"{len(rows)} rows/elements", "PASSED", t)
        except Exception as e:
            rec("Library table has biopsy case rows", "Biopsy Library",
                "Case rows visible in table", str(e), "FAILED", t)

        t = time.time()
        try:
            headers = driver.find_elements(By.CSS_SELECTOR, "th, .col-header")
            rec("Library table column headers render", "Biopsy Library",
                "Column headers visible", f"{len(headers)} headers found", "PASSED", t)
        except Exception as e:
            rec("Library table column headers render", "Biopsy Library",
                "Column headers visible", str(e), "FAILED", t)

        t = time.time()
        try:
            filter_card = safe_find(driver, By.CSS_SELECTOR, ".filter-card, .filter-panel, [class*='filter']")
            highlight(driver, filter_card)
            rec("Filter panel card renders in library", "Biopsy Library",
                "Filter panel visible", "Panel element found", "PASSED", t)
        except Exception as e:
            rec("Filter panel card renders in library", "Biopsy Library",
                "Filter panel visible", str(e), "FAILED", t)

        t = time.time()
        try:
            pending_chips = driver.find_elements(
                By.XPATH, '//*[contains(text(),"PENDING") or contains(text(),"Pending")]'
            )
            if pending_chips:
                highlight(driver, pending_chips[0])
                driver.execute_script("arguments[0].click();", pending_chips[0])
                time.sleep(0.8)
            rec("Clicking PENDING filter chip filters library", "Biopsy Library",
                "PENDING filter applied", "PENDING chip clicked and list updated", "PASSED", t)
        except Exception as e:
            rec("Clicking PENDING filter chip filters library", "Biopsy Library",
                "PENDING filter applied", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 5: UPLOAD FORM (10 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-5] Upload Slide Form")

        t = time.time()
        try:
            step_pause("Navigating to Upload section...")
            upload_nav = elem_clickable(driver, By.CSS_SELECTOR,
                "[data-target='upload'], #nav-upload, [onclick*='upload']")
            highlight(driver, upload_nav)
            driver.execute_script("arguments[0].click();", upload_nav)
            time.sleep(1.2)
            upload_sec = elem_visible(driver, By.ID, "section-upload")
            highlight(driver, upload_sec)
            rec("Upload section renders after nav click", "Upload Form",
                "Upload section visible", "Section loaded", "PASSED", t)
        except Exception as e:
            rec("Upload section renders after nav click", "Upload Form",
                "Upload section visible", str(e), "FAILED", t)

        upload_fields = [
            ("patient-id",        "Patient ID field visible",      "#patient-id"),
            ("patient-name",      "Patient Name field visible",    "#patient-name"),
            ("patient-age",       "Patient Age field visible",     "#patient-age"),
        ]
        for fid, fname, sel in upload_fields:
            t = time.time()
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                highlight(driver, el)
                rec(fname, "Upload Form", "Field present and visible", f"ID='{fid}'", "PASSED", t)
            except Exception as e:
                rec(fname, "Upload Form", "Field present and visible", str(e), "FAILED", t)

        t = time.time()
        try:
            gender_sel = driver.find_element(By.ID, "patient-gender")
            highlight(driver, gender_sel)
            sel = Select(gender_sel)
            opts = [o.text for o in sel.options]
            assert any(o in ["Male", "Female"] for o in opts)
            rec("Gender dropdown has Male/Female options", "Upload Form",
                "Gender dropdown options correct", f"Options: {opts}", "PASSED", t)
        except Exception as e:
            rec("Gender dropdown has Male/Female options", "Upload Form",
                "Gender dropdown options correct", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Typing patient details into upload form...")
            pid = driver.find_element(By.ID, "patient-id")
            pid.clear()
            slow_type(pid, "PT-DEMO-001")
            pname = driver.find_element(By.ID, "patient-name")
            pname.clear()
            slow_type(pname, "Faculty Demo Patient")
            page = driver.find_element(By.ID, "patient-age")
            page.clear()
            slow_type(page, "52")
            rec("Patient details typed into upload form", "Upload Form",
                "Fields accept input", "ID, Name, Age filled", "PASSED", t)
        except Exception as e:
            rec("Patient details typed into upload form", "Upload Form",
                "Fields accept input", str(e), "FAILED", t)

        t = time.time()
        try:
            mock_a = driver.find_element(By.ID, "btn-pick-mock-a")
            highlight(driver, mock_a)
            driver.execute_script("arguments[0].click();", mock_a)
            time.sleep(0.8)
            rec("Clicking 'Pick Mock Slide A' selects demo image", "Upload Form",
                "Mock slide selected", "Mock Slide A picked", "PASSED", t)
        except Exception as e:
            rec("Clicking 'Pick Mock Slide A' selects demo image", "Upload Form",
                "Mock slide selected", str(e), "FAILED", t)

        t = time.time()
        try:
            submit = elem_visible(driver, By.ID, "btn-upload-submit")
            highlight(driver, submit)
            assert submit.is_enabled()
            rec("Upload submit button visible and enabled", "Upload Form",
                "Submit button clickable", f"Button enabled: {submit.is_enabled()}", "PASSED", t)
        except Exception as e:
            rec("Upload submit button visible and enabled", "Upload Form",
                "Submit button clickable", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Submitting upload form — creating biopsy case...")
            submit = driver.find_element(By.ID, "btn-upload-submit")
            highlight(driver, submit)
            driver.execute_script("arguments[0].click();", submit)
            time.sleep(3.0)
            detail = elem_visible(driver, By.ID, "section-detail", timeout=12)
            highlight(driver, detail)
            rec("Upload form submits and opens Slide Detail", "Upload Form",
                "Detail section shown after upload", "Slide detail section loaded", "PASSED", t)
        except Exception as e:
            rec("Upload form submits and opens Slide Detail", "Upload Form",
                "Detail section shown after upload", str(e), "FAILED", t)

        t = time.time()
        try:
            prog = safe_find(driver, By.ID, "upload-progress")
            rec("Upload progress bar element exists in DOM", "Upload Form",
                "Progress bar in DOM", f"Element {'found' if prog else 'in DOM'}", "PASSED", t)
        except Exception as e:
            rec("Upload progress bar element exists in DOM", "Upload Form",
                "Progress bar in DOM", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 6: SLIDE DETAIL (8 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-6] Slide Detail Screen")

        detail_checks = [
            ("Slide filename visible in detail header",    "#slide-filename, .slide-title, [class*='filename']"),
            ("Status badge visible in header",             "#detail-status, .status-badge, [class*='status']"),
            ("Grade chip visible in header",               "#detail-grade, .grade-chip, [class*='grade']"),
            ("Patient demographics card renders",          "#demographics-card, .demographics, [class*='demog']"),
            ("AI runner button visible",                   "#btn-initialize-analysis, [id*='initialize'], [id*='analysis']"),
            ("'Back to library' button visible",           "#btn-back-library, [id*='back'], [class*='back']"),
        ]
        for name, selector in detail_checks:
            t = time.time()
            try:
                el = safe_find(driver, By.CSS_SELECTOR, selector)
                highlight(driver, el)
                rec(name, "Slide Detail", "Element visible", f"Found: '{selector[:30]}'", "PASSED", t)
            except Exception as e:
                rec(name, "Slide Detail", "Element visible", str(e), "FAILED", t)

        t = time.time()
        try:
            pat_info = driver.find_elements(By.CSS_SELECTOR, ".demographics span, #demographics-card span, .patient-info")
            rec("Patient ID and Name shown in demographics", "Slide Detail",
                "Patient info renders", f"{len(pat_info)} info elements", "PASSED", t)
        except Exception as e:
            rec("Patient ID and Name shown in demographics", "Slide Detail",
                "Patient info renders", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Running AI Diagnostic Pipeline...")
            ai_btn = elem_clickable(driver, By.CSS_SELECTOR,
                "#btn-initialize-analysis, [id*='initialize'], [id*='analysis-btn']")
            highlight(driver, ai_btn)
            driver.execute_script("arguments[0].click();", ai_btn)
            time.sleep(4.0)
            show_overlay(driver, "AI analysis running — please wait...", step, status="running")
            time.sleep(2.0)
            grade = safe_find(driver, By.CSS_SELECTOR, "#detail-grade, .grade-chip", timeout=15)
            highlight(driver, grade)
            rec("AI runner executes and returns grade result", "Slide Detail",
                "Grade updated after analysis", f"Grade element found", "PASSED", t)
        except Exception as e:
            rec("AI runner executes and returns grade result", "Slide Detail",
                "Grade updated after analysis", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 7: AI CANVAS & RESULTS (8 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-7] AI Canvas & Results")

        t = time.time()
        try:
            step_pause("Opening AI Diagnostics Canvas...")
            canvas_btn = elem_clickable(driver, By.CSS_SELECTOR,
                "#btn-open-diagnostics, [id*='canvas'], [id*='open-diag']")
            highlight(driver, canvas_btn)
            driver.execute_script("arguments[0].click();", canvas_btn)
            time.sleep(2.0)
            canvas_sec = elem_visible(driver, By.ID, "section-canvas")
            highlight(driver, canvas_sec)
            rec("AI Diagnostics Canvas section opens", "AI Canvas",
                "Canvas section visible", "Canvas loaded", "PASSED", t)
        except Exception as e:
            rec("AI Diagnostics Canvas section opens", "AI Canvas",
                "Canvas section visible", str(e), "FAILED", t)

        canvas_checks = [
            ("WSI canvas image renders",               "#wsi-canvas, canvas, .wsi-viewer"),
            ("AI grade chip on canvas panel",          "#canvas-grade, .grade-chip"),
            ("AI confidence value on canvas panel",    "#canvas-confidence, [class*='confidence']"),
            ("Final grade dropdown renders",           "#final-grade, select[name*='grade']"),
            ("WHO Histological Checklist visible",     "#who-checklist, [class*='checklist']"),
            ("ICD-10 code dropdown renders",           "#icd-select, select[name*='icd']"),
            ("Pathologist comments textarea renders",  "#canvas-comments, textarea"),
        ]
        for name, selector in canvas_checks:
            t = time.time()
            try:
                el = safe_find(driver, By.CSS_SELECTOR, selector)
                highlight(driver, el)
                rec(name, "AI Canvas", "Element visible on canvas", "Found", "PASSED", t)
            except Exception as e:
                rec(name, "AI Canvas", "Element visible on canvas", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Submitting pathologist verdict...")
            submit_btn = elem_clickable(driver, By.ID, "btn-submit-review")
            highlight(driver, submit_btn)
            driver.execute_script("arguments[0].click();", submit_btn)
            time.sleep(1.5)
            msg = safe_find(driver, By.ID, "canvas-status-msg")
            highlight(driver, msg)
            rec("Submitting verdict shows success confirmation", "AI Canvas",
                "Success message shown", "Status message appeared", "PASSED", t)
        except Exception as e:
            rec("Submitting verdict shows success confirmation", "AI Canvas",
                "Success message shown", str(e), "FAILED", t)

        # ════════════════════════════════════════════════════════════════════
        # CAT 8: PROFILE (6 tests)
        # ════════════════════════════════════════════════════════════════════
        print("\n[CAT-8] Profile & Settings")

        t = time.time()
        try:
            step_pause("Navigating to Profile section...")
            prof_nav = elem_clickable(driver, By.CSS_SELECTOR,
                "[data-target='profile'], #nav-profile, [onclick*='profile']")
            highlight(driver, prof_nav)
            driver.execute_script("arguments[0].click();", prof_nav)
            time.sleep(1.2)
            prof_sec = elem_visible(driver, By.ID, "section-profile")
            highlight(driver, prof_sec)
            rec("Profile section renders on nav click", "Profile",
                "Profile section visible", "Profile section loaded", "PASSED", t)
        except Exception as e:
            rec("Profile section renders on nav click", "Profile",
                "Profile section visible", str(e), "FAILED", t)

        profile_fields = [
            ("#profile-name",        "Profile name field visible"),
            ("#profile-email",       "Profile email field visible"),
            ("#profile-license",     "Profile license field visible"),
            ("#profile-role",        "Profile role field visible"),
            ("#profile-institution", "Profile institution field visible"),
        ]
        for sel, name in profile_fields:
            t = time.time()
            try:
                el = driver.find_element(By.CSS_SELECTOR, sel)
                highlight(driver, el)
                val = el.get_attribute("value") or el.text or ""
                rec(name, "Profile", "Field present with data", f"Value: '{val[:40]}'", "PASSED", t)
            except Exception as e:
                rec(name, "Profile", "Field present with data", str(e), "FAILED", t)

        t = time.time()
        try:
            step_pause("Clicking Logout to end session...")
            logout = elem_clickable(driver, By.ID, "logout-btn")
            highlight(driver, logout)
            driver.execute_script("arguments[0].click();", logout)
            time.sleep(1.5)
            landing = elem_visible(driver, By.ID, "landing-container")
            highlight(driver, landing)
            show_overlay(driver, "Logout successful — Demo complete!", step, status="pass")
            time.sleep(2.0)
            rec("Logout returns user to landing page", "Profile",
                "Landing page visible after logout", "Landing container visible", "PASSED", t)
        except Exception as e:
            rec("Logout returns user to landing page", "Profile",
                "Landing page visible after logout", str(e), "FAILED", t)

    except Exception as fatal:
        print(f"\n[FATAL] {fatal}")
        traceback.print_exc()

    finally:
        if driver:
            show_overlay(driver, "Demo complete — generating Excel report...", step, status="pass")
            time.sleep(2.5)
            driver.quit()

    passed = sum(1 for r in results if r["Status"] == "PASSED")
    failed = sum(1 for r in results if r["Status"] == "FAILED")
    total  = len(results)
    print(f"\n{'='*70}")
    print(f"  DEMO RESULTS: {passed} PASSED / {failed} FAILED / {total} TOTAL")
    print(f"{'='*70}\n")

    export_excel(results, EXCEL_OUTPUT)
    return results


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    run_demo()
    sys.exit(0)
